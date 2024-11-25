import tkinter as tk
from tkinter import ttk
import subprocess
import pyodbc
import customtkinter
from tkinter import PhotoImage
import customtkinter as ctk
from tkinter import messagebox, filedialog
# import pandas as pd
import arrow
from datetime import datetime
from PIL import Image, ImageTk  # Import thư viện để làm việc với ảnh
from admin_login import create_login_dialog
# from pip._vendor.urllib3.util.timeout import current_time
import openpyxl
from openpyxl.styles import Font, Alignment
# from tkinter import simpledialog
import re
from datetime import time
import pickle
import cv2
import bcrypt
from insightface.app import FaceAnalysis
import os
import sys
import warnings
warnings.simplefilter("ignore", FutureWarning)
from customtkinter import CTkImage
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side



customtkinter.set_appearance_mode("light")
customtkinter.set_default_color_theme("blue")

def close_dialog(dialog):
    dialog.destroy()

def is_valid_email(email):
    # Biểu thức chính quy để kiểm tra định dạng email
    pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.match(pattern, email) is not None

def update_time():
    now = datetime.now()
    time_label.config(text=now.strftime("%H:%M:%S"))
    date_label.config(text=now.strftime("%d/%m/%Y"))
    root.after(1000, update_time)

# def center_dialog2(window):
#         window.update_idletasks()  # Cập nhật kích thước
#         width = window.winfo_width()
#         height = window.winfo_height()
#         screen_width = window.winfo_screenwidth()
#         screen_height = window.winfo_screenheight()
#         x = (screen_width // 2) - (width // 2)
#         y = (screen_height // 2) - (height // 2)
#         window.geometry(f"{width}x{height}+{x}+{y}")

def center_dialog(dialog):
    dialog.update_idletasks()  # Cập nhật các tác vụ chờ
    x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
    y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
    dialog.geometry(f"+{x}+{y}")  # Đặt vị trí cửa sổ

def show_user_management():
    # Ẩn menu khác
    attendance_frame.pack_forget()
    catalog_frame.pack_forget()
    user_management_frame.pack(fill='both', expand=True)

def show_attendance_management():
    # Ẩn menu khác
    user_management_frame.pack_forget()
    catalog_frame.pack_forget()
    attendance_frame.pack(fill='both', expand=True)


def show_catalog_management():
    user_management_frame.pack_forget()
    attendance_frame.pack_forget()
    catalog_frame.pack(fill='both', expand=True)
    
def get_userid():
    with open("userid.txt", "r") as f:
        return int(f.read().strip())
    
def save_db_config(server_name, database_name):
    with open("db_config.txt", "w") as config_file:
        config_file.write(f"{server_name}\n{database_name}")

# Hàm tải thông tin server và database từ tệp (nếu có)
def load_db_config():
    try:
        with open("db_config.txt", "r") as config_file:
            lines = config_file.readlines()
            if len(lines) >= 2:
                return lines[0].strip(), lines[1].strip()
    except FileNotFoundError:
        pass
    return None, None



def check_db_connection():
    """Tạo dialog nhập server name và database name để kiểm tra kết nối."""
    dialog = customtkinter.CTkToplevel(root)
    dialog.title("Kiểm tra kết nối DB")
    center_dialog(dialog)
    dialog.geometry("330x250")
    
    dialog.grab_set()

    # Tải thông tin từ tệp nếu có
    saved_server, saved_database = load_db_config()

    # Label và Entry cho Server Name
    customtkinter.CTkLabel(dialog, text="Server Name:").place(relx=0.1, rely=0.1)
    server_entry = customtkinter.CTkEntry(dialog, width=250)
    server_entry.place(relx=0.1, rely=0.2)
    if saved_server:
        server_entry.insert(0, saved_server)  # Hiển thị thông tin đã lưu

    # Label và Entry cho Database Name
    customtkinter.CTkLabel(dialog, text="Database Name:").place(relx=0.1, rely=0.4)
    database_entry = customtkinter.CTkEntry(dialog, width=250)
    database_entry.place(relx=0.1, rely=0.5)
    if saved_database:
        database_entry.insert(0, saved_database)  # Hiển thị thông tin đã lưu

    # Label để hiển thị thông báo
    result_label = customtkinter.CTkLabel(dialog, text="")
    result_label.pack(pady=10)

    # Hàm kiểm tra kết nối
    def connect_db():
        server_name = server_entry.get()
        database_name = database_entry.get()

        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )
            info_label.configure(text="Kết nối thành công!")
            save_db_config(server_name, database_name)  # Lưu thông tin sau khi kết nối thành công

        except pyodbc.Error as ex:
            print('Kết nối tới DB thất bại!', ex)
            info_label.configure(text="Kết nối tới DB thất bại!")

    connect_button = customtkinter.CTkButton(dialog, text="Kết nối DB",
                                             command=connect_db,
                                             fg_color="blue")
    connect_button.place(relx=0.1, rely=0.7)

    info_label = customtkinter.CTkLabel(dialog, text="")
    info_label.place(relx=0.1, rely=0.83)

# Hàm kiểm tra sự tồn tại của user trong face_data.pkl
def check_user_in_face_data(user_id):
    try:
        # Mở và đọc dữ liệu từ face_data.pkl
        with open('face_data.pkl', 'rb') as f:
            face_data = pickle.load(f)
        
        # Kiểm tra kiểu dữ liệu của face_data
        # print(f"Kiểu dữ liệu của face_data: {type(face_data)}")
        
        # Kiểm tra xem face_data có phải là dictionary không
        if isinstance(face_data, dict):
            # Kiểm tra xem user_id có tồn tại trong dictionary hay không
            if str(user_id) in face_data:  # Chú ý đổi sang kiểu string nếu key là string
                return True  # Có khuôn mặt đã được đăng ký
            else:
                return False  # Chưa có khuôn mặt được đăng ký
        else:
            print("Dữ liệu không phải là dictionary!")
            return False
    except FileNotFoundError:
        print("File face_data.pkl không tồn tại.")
        return False
    except Exception as e:
        print(f"Đã xảy ra lỗi khi kiểm tra dữ liệu trong face_data.pkl: {e}")
        return False


# Hàm để load dữ liệu user và thêm vào treeview
def load_user_data():
    # Kết nối đến cơ sở dữ liệu
    server_name, database_name = load_db_config()
    if server_name and database_name:
        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )

            cursor = connection.cursor()
            cursor.execute("SELECT UserID, Name, Role, job_position, Username, Email, Note FROM user_manager")  # Lấy 7 cột
            rows = cursor.fetchall()

            # Xóa dữ liệu cũ trong Treeview
            for item in user_treeview.get_children():
                user_treeview.delete(item)
            

            
            # Thêm dữ liệu mới vào Treeview
            for row in rows:
                user_id = row[0]
                is_registered = check_user_in_face_data(user_id)
                
                # Chọn icon dựa trên việc đăng ký khuôn mặt
                check_value = "Đã thêm" if is_registered else "Chưa thêm"
                
                user_treeview.insert('', 'end', values=(row[0], row[1], row[2], row[3], row[4], row[5], row[6]),
                                     tags=('centered',))
                # Thêm icon vào cột 'Check' (Cột 7)
                # user_treeview.set(user_treeview.get_children()[-1], column='Check', value=check_icon)
                user_treeview.set(user_treeview.get_children()[-1], column='Check', value=check_value)
            cursor.close()
            connection.close()
        except pyodbc.Error as ex:
            messagebox.showerror("Lỗi", "Lỗi tải dữ liệu người dùng!")
            # print("Error loading user data:", ex)
    else:
        messagebox.showerror("Lỗi", "Servername hoặc database bị thiếu!")
        # print("Server name or database name is missing.")



def dialog_search_data():
    search_dialog = tk.Toplevel(root)  # Tạo cửa sổ mới
    search_dialog.title("Tìm Kiếm")
    search_dialog.geometry("400x290")
    search_dialog.resizable(False, False)
    search_dialog.grab_set()

    center_dialog(search_dialog)
    # Các trường nhập liệu
    labels = ['UserID', 'Tên', 'Vai trò','Vị trí', 'Tên đăng nhập', 'Email', 'Ghi chú']
    entries = []

    for label in labels:
        row = tk.Frame(search_dialog)
        lab = tk.Label(row, text=label, width=15, anchor='w')  # Căn trái
        ent = tk.Entry(row, width=30)
        lab.pack(side=tk.LEFT, padx=(10, 5), pady=(2, 0))  # Khoảng cách 10px bên trái, 5px bên phải
        ent.pack(side=tk.RIGHT, padx=(0, 10), pady=(2, 0))  # Khoảng cách 10px bên phải
        row.pack(pady=5)
        entries.append(ent)

    # Nút tìm kiếm
    search_button = tk.Button(search_dialog, text="Tìm Kiếm", command=lambda: search_data2(entries, search_dialog), width=20)
    search_button.pack(pady=10)

    def user_data_search():
        # Kết nối đến cơ sở dữ liệu
        server_name, database_name = load_db_config()
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )

                cursor = connection.cursor()
                cursor.execute("SELECT UserID, Name, Role,job_position, Username, Email, Note FROM user_manager")  # Lấy 6 cột
                rows = cursor.fetchall()
                cursor.close()
                connection.close()
                

                return rows  # Trả về dữ liệu để sử dụng trong các hàm khác
            except pyodbc.Error as ex:
                messagebox.showerror("Lỗi", "Lỗi tải dữ liệu người dùng")
                # print("Error loading user data:", ex)
                return []  # Trả về danh sách rỗng nếu có lỗi
        else:
            messagebox.showerror("Lỗi", "Servername hoặc database bị thiếu!")
            # print("Server name or database name is missing.")
            return []  # Trả về danh sách rỗng nếu không có thông tin kết nối

    def search_data2(entries, search_dialog):
        search_terms = [entry.get().lower() for entry in entries]  
        for item in user_treeview.get_children():
            user_treeview.delete(item)

        all_data = user_data_search()  # Sử dụng load_user_data() để lấy dữ liệu từ SQL
        found_records = 0  # Biến đánh dấu có tìm thấy bản ghi hay không

        for record in all_data:
            conditions = [
                (not search_terms[0] or str(record[0]) == search_terms[0]),  # ID
                (not search_terms[1] or search_terms[1] in record[1].lower()),  # Tên
                (not search_terms[2] or search_terms[2] in record[2].lower()),  # Vai trò
                (not search_terms[3] or search_terms[3] in record[3].lower()),  # vi tri
                (not search_terms[4] or search_terms[4].lower() in (record[4] or "").lower()),  # ten dang nhap
                (not search_terms[5] or search_terms[5].lower() in (record[5] or "").lower()),#email
                (not search_terms[6] or search_terms[6].lower() in (record[6] or "").lower()),# Ghi chú
            ]
            
            if all(conditions):  # Nếu tất cả các điều kiện thỏa mãn
                user_treeview.insert("", "end", values=(record[0], record[1], record[2], record[3], record[4], record[5],record[6]))
                found_records += 1  # Đánh dấu là đã tìm thấy bản ghi

        if found_records == 0:  # Nếu không tìm thấy bản ghi nào
            show_no_record_found_dialog()
        else:
            search_dialog.destroy()  # Chỉ đóng dialog tìm kiếm nếu có bản ghi được tìm thấy
            show_found_records_info(found_records)  # Hiển thị số lượng bản ghi tìm thấy

    def show_no_record_found_dialog():
        no_record_dialog = tk.Toplevel(root)
        
        no_record_dialog.title("Thông báo")
        center_dialog(no_record_dialog)
        no_record_dialog.geometry("300x150")  # Thiết lập kích thước cố định trước
        # Căn giữa sau khi thiết lập kích thước
        no_record_dialog.grab_set()  # Đảm bảo dialog này nhận tất cả sự chú ý
        
        message = tk.Label(no_record_dialog, text="Không tìm thấy bản ghi phù hợp!", padx=10, pady=10)
        message.pack()

        ok_button = tk.Button(no_record_dialog, text="OK", command=no_record_dialog.destroy,width=15, height=2,bg="green")
        ok_button.pack(pady=10, padx=10)

        no_record_dialog.focus_set()  # Đặt focus cho dialog thông báo
        center_dialog(no_record_dialog)

    def show_found_records_info(count):
        info_dialog = tk.Toplevel(root)
        info_dialog.title("Thông Báo")
        info_dialog.geometry("300x150")
        info_dialog.resizable(False, False)
        info_dialog.grab_set()

        message = tk.Label(info_dialog, text=f"Đã tìm thấy {count} bản ghi!", padx=10, pady=10)
        message.pack()

        ok_button = tk.Button(info_dialog, text="OK", command=info_dialog.destroy, width=15, height=2, bg="blue")
        ok_button.pack(pady=10)

        info_dialog.focus_set()
        center_dialog(info_dialog)

def get_next_user_id():
    """Lấy ID người dùng tiếp theo từ cơ sở dữ liệu theo số còn thiếu nhỏ nhất."""
    server_name, database_name = load_db_config()
    if server_name and database_name:
        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )
            cursor = connection.cursor()

            # Lấy tất cả các UserID hiện có và sắp xếp tăng dần
            cursor.execute("SELECT UserID FROM user_manager ORDER BY UserID ASC")
            user_ids = [row[0] for row in cursor.fetchall()]
            cursor.close()
            connection.close()

            # Tìm số thiếu nhỏ nhất
            max_id = user_ids[-1] if user_ids else 0
            for i in range(1, max_id + 1):
                if i not in user_ids:
                    return i  # Trả về số thiếu nhỏ nhất
            
            # Nếu không thiếu, trả về MaxID + 1
            return max_id + 1

        except pyodbc.Error as ex:
            messagebox.showerror("Lỗi", "Lỗi tìm kiếm ID người dùng tiếp theo!")
            # print("Error fetching next user ID:", ex)
            return 1
    return 1

def get_job_positions():
    # Kết nối đến cơ sở dữ liệu và lấy danh sách các vị trí
    server_name, database_name = load_db_config()
    if server_name and database_name:
        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )
            cursor = connection.cursor()
            cursor.execute("SELECT DISTINCT job_position FROM job_position")  # Lấy các vị trí duy nhất
            job_positions = [row.job_position for row in cursor.fetchall()]  # Tạo danh sách các vị trí
            cursor.close()
            connection.close()
            return job_positions
        except pyodbc.Error as ex:
            messagebox.showerror("Lỗi", "Lỗi tìm kiếm vị trí việc làm!")
            # print("Error fetching job positions:", ex)
            return []
    else:
        messagebox.showerror("Lỗi", "Servername hoặc database bị thiếu!")
        return []


def is_valid_email_format(email):
    # Kiểm tra định dạng email bằng regex
    email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(email_regex, email) is not None

insightface_path = os.path.join(os.getcwd(), 'insightface')  
if insightface_path not in sys.path:
    sys.path.append(insightface_path)
    
face_recognition_app = FaceAnalysis()
face_recognition_app.prepare(ctx_id=0)


face_data = {}
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
def detect_face(frame):
    global face_recognition_app
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    
    # Phát hiện khuôn mặt sử dụng Haar Cascade
    faces = face_cascade.detectMultiScale(gray, 1.1, 4)

    if len(faces) > 0:
        for (x, y, w, h) in faces:
            # Vẽ khung vuông quanh khuôn mặt
            cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
        
        # Lấy đặc trưng khuôn mặt
        face_embedding = face_recognition_app.get(frame)
        if len(face_embedding) > 0:
            return face_embedding[0].embedding, faces
    return None, None


def add_face_user():
    selected_item = user_treeview.selection()  # Lấy item đang được chọn trong Treeview

    if not selected_item:  # Nếu không có bản ghi nào được chọn
        messagebox.showerror("Lỗi", "Vui lòng chọn một bản ghi người dùng để thêm khuôn mặt.")
        return  # Dừng hàm nếu không có bản ghi được chọn
    
    # Lấy dữ liệu UserID và Name từ bản ghi được chọn
    user_face_id = user_treeview.item(selected_item[0], 'values')[0]  # Cột 1: UserID
    name = user_treeview.item(selected_item[0], 'values')[1]  # Cột 2: Name
    
    # Tạo cửa sổ con để hiển thị thông tin và video
    add_user_window = tk.Toplevel()
    add_user_window.title("Thêm Khuôn Mặt")
    add_user_window.geometry("640x520")
    center_dialog(add_user_window)
    add_user_window.grab_set()
    
    # Phần 1: Hiển thị thông tin người dùng (UserID và Name)
    user_label = tk.Label(add_user_window, text=f"UserID: {user_face_id}")
    user_label.place(x=20, y=10)
    
    name_label = tk.Label(add_user_window, text=f"Tên: {name}")
    name_label.place(x=320, y=10)

    # Phần 2: Video label để hiển thị camera
    video_label = tk.Label(add_user_window)
    video_label.place(x=20, y=50)

    # Mở camera và lấy khung hình
    cap = cv2.VideoCapture(0)

    # Hàm cập nhật và hiển thị frame từ camera
    def update_frame():
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Lỗi", "Không thể mở camera.")
            return
        frame = cv2.flip(frame, 1)
        frame = cv2.resize(frame, (600, 400))
        
        # Phát hiện khuôn mặt và vẽ lên frame
        face_embedding, faces = detect_face(frame)

        # Hiển thị video lên giao diện
        img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(img)
        img = img.resize((600, 400))
        img = ImageTk.PhotoImage(img)
        video_label.configure(image=img)
        video_label.image = img

        # Gọi lại update_frame sau mỗi 10ms để cập nhật liên tục
        add_user_window.after(10, update_frame)

    # Hàm lưu dữ liệu khi nút "Bắt Đầu" được bấm
    def start_recognition():
        face_embedding, faces = detect_face(cap.read()[1])  # Lấy một frame và nhận diện khuôn mặt
        if face_embedding is not None:
            save_data(face_embedding, user_face_id, name)

    # Hàm lưu khuôn mặt vào dữ liệu (thêm mới, không ghi đè)
    def save_data(face_embedding, user_face_id, name):
        try:
            # Đọc dữ liệu cũ từ file nếu có
            with open("face_data.pkl", "rb") as f:
                face_data = pickle.load(f)
        except (FileNotFoundError, EOFError):
            # Nếu file không tồn tại hoặc rỗng, tạo dictionary rỗng
            face_data = {}

        # Cập nhật thông tin khuôn mặt vào dữ liệu cũ
        face_data[user_face_id] = {
            'userid': user_face_id,
            'name': name,
            'embedding': face_embedding.tolist()  # Chuyển numpy array thành list
        }

        # Ghi lại tất cả dữ liệu vào file
        with open("face_data.pkl", "wb") as f:
            pickle.dump(face_data, f)

        messagebox.showinfo("Thông báo", f"Khuôn mặt của {name} đã được lưu thành công!")
        add_user_window.destroy()  # Đóng cửa sổ
        cap.release()  # Giải phóng camera
        cv2.destroyAllWindows()
    
    # Nút "Bắt Đầu" để bắt đầu quá trình nhận diện và lưu dữ liệu
    start_button = tk.Button(add_user_window, text="Bắt Đầu", command=start_recognition,width=20)
    start_button.place(x=250, y=470)

    # Giải phóng camera sau khi hoàn thành
    def on_close():
        cap.release()
        cv2.destroyAllWindows()
        add_user_window.destroy()

    add_user_window.protocol("WM_DELETE_WINDOW", on_close)

    # Bắt đầu nhận diện video khi cửa sổ được mở
    update_frame()
    
def add_user():
    add_dialog = customtkinter.CTkToplevel(root)
    add_dialog.title("Thêm Người Dùng")
    add_dialog.geometry("400x500")  # Tăng kích thước một chút để chứa thêm ô nhập vị trí
    
    add_dialog.grab_set()
    next_id = get_next_user_id()
    customtkinter.CTkLabel(add_dialog, text=f"ID Người Dùng Mới: {next_id}").place(relx=0.1, rely=0.05)

    customtkinter.CTkLabel(add_dialog, text="Tên:").place(relx=0.1, rely=0.1)
    name_entry = customtkinter.CTkEntry(add_dialog, width=310)
    name_entry.place(relx=0.1, rely=0.15)
        
    # Lấy danh sách các vị trí từ bảng job_position

    job_positions = get_job_positions()  # Lấy danh sách các vị trí từ cơ sở dữ liệu

    customtkinter.CTkLabel(add_dialog, text="Vị trí:").place(relx=0.1, rely=0.27)
    position_var = tk.StringVar()  # Biến để lưu lựa chọn vị trí
    position_combobox = ttk.Combobox(add_dialog, textvariable=position_var, values=job_positions)
    position_combobox.place(relx=0.3, rely=0.27)


    customtkinter.CTkLabel(add_dialog, text="Vai trò:").place(relx=0.1, rely=0.38)
    role_var = tk.StringVar(value="user")
    role_combobox = ttk.Combobox(add_dialog, textvariable=role_var, values=["admin", "user"])
    role_combobox.place(relx=0.3, rely=0.38)
    



    # Các trường nhập liệu cho admin
    username_label = customtkinter.CTkLabel(add_dialog, text="Tên Đăng Nhập:")
    password_label = customtkinter.CTkLabel(add_dialog, text="Mật Khẩu:")
    email_label = customtkinter.CTkLabel(add_dialog, text="Email:")
    
    username_entry = customtkinter.CTkEntry(add_dialog, width=310)
    password_entry = customtkinter.CTkEntry(add_dialog, width=310, show='*')
    email_entry = customtkinter.CTkEntry(add_dialog, width=310)

    # Label và Entry cho Ghi Chú
    note_label = customtkinter.CTkLabel(add_dialog, text="Ghi Chú:")
    note_entry = customtkinter.CTkEntry(add_dialog, width=310)
    note_entry.place(relx=0.1, rely=0.8)

    def update_fields(event):
        if role_var.get() == "admin":
            add_dialog.geometry("400x650")
            username_label.place(relx=0.1, rely=0.45)
            username_entry.place(relx=0.1, rely=0.5)
            password_label.place(relx=0.1, rely=0.55)
            password_entry.place(relx=0.1, rely=0.6)
            email_label.place(relx=0.1, rely=0.65)
            email_entry.place(relx=0.1, rely=0.7)
            note_label.place(relx=0.1, rely=0.75)
            note_entry.place(relx=0.1, rely=0.8)
            center_dialog(add_dialog)
        else:
            add_dialog.geometry("400x500")
            username_label.place_forget()
            username_entry.place_forget()
            password_label.place_forget()
            password_entry.place_forget()
            email_label.place_forget()
            email_entry.place_forget()
            note_label.place(relx=0.1, rely=0.45)
            note_entry.place(relx=0.1, rely=0.5)
            center_dialog(add_dialog)

    role_combobox.bind("<<ComboboxSelected>>", update_fields)

    result_label = customtkinter.CTkLabel(add_dialog, text="")
    result_label.place(relx=0.1, rely=0.85)

    def save_user():
        user_id = get_next_user_id()  # Lấy ID người dùng mới
        name = name_entry.get()
        role = role_var.get()
        position = position_var.get()  # Lấy vị trí người dùng đã chọn
        username = username_entry.get() if role == "admin" else None
        password = password_entry.get() if role == "admin" else None
        email = email_entry.get() if role == "admin" else None
        note = note_entry.get()

        server_name, database_name = load_db_config()
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )
                cursor = connection.cursor()

                # Kiểm tra xem username đã tồn tại chưa
                if role == "admin" and username:
                    cursor.execute("SELECT COUNT(*) FROM user_manager WHERE Username = ?", (username,))
                    result = cursor.fetchone()
                    if result[0] > 0:
                        result_label.configure(text="Tên đăng nhập đã tồn tại, vui lòng chọn tên khác.")
                        cursor.close()
                        connection.close()
                        return

                # Kiểm tra định dạng email
                if role == "admin" and email:
                    if not is_valid_email_format(email):
                        result_label.configure(text="Email không đúng định dạng, vui lòng nhập lại.")
                        cursor.close()
                        connection.close()
                        return

                    # Kiểm tra xem email đã tồn tại chưa
                    cursor.execute("SELECT COUNT(*) FROM user_manager WHERE Email = ?", (email,))
                    result = cursor.fetchone()
                    if result[0] > 0:
                        result_label.configure(text="Email đã tồn tại, vui lòng chọn email khác.")
                        cursor.close()
                        connection.close()
                        return
                if role == "admin" and password:
                    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())  # Mã hóa mật khẩu
                    print(hashed_password)
                hashed_password_str = hashed_password.decode('utf-8')

                # Thêm người dùng nếu các điều kiện đều thỏa mãn
                cursor.execute(
                    "INSERT INTO user_manager (UserID, Name, Role, job_position, Username, Password, Email, Note) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (user_id, name, role, position, username, hashed_password_str, email, note)
                )
                connection.commit()
                cursor.close()
                connection.close()
                result_label.configure(text="Thêm người dùng thành công!")
                load_user_data()  # Tải lại dữ liệu người dùng

            except pyodbc.Error as ex:
                # messagebox.showerror("Lỗi", "Lỗi lưu dữ liệu người dùng!")
                # print("Error saving user data:", ex)
                result_label.configure(text="Lỗi khi thêm người dùng.")
        else:
            # print("Server name or database name is missing.")
            result_label.configure(text="Lỗi kết nối cơ sở dữ liệu.")

    save_button = customtkinter.CTkButton(add_dialog, text="Lưu", command=save_user, fg_color="green")
    save_button.place(relx=0.35, rely=0.9)

    update_fields(None)

def edit_user():
    # Kiểm tra xem có chọn một bản ghi nào chưa
    selected_item = user_treeview.selection()
    if not selected_item:
        messagebox.showerror("Lỗi", "Vui lòng chọn một bản ghi để sửa!")
        # print("Vui lòng chọn một bản ghi để sửa.")
        return

    # Lấy ID của bản ghi đã chọn
    selected_id = user_treeview.item(selected_item, 'values')[0]

    # Kết nối đến cơ sở dữ liệu và lấy dữ liệu của bản ghi đã chọn
    server_name, database_name = load_db_config()
    if server_name and database_name:
        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )
            cursor = connection.cursor()
            cursor.execute("SELECT UserID, Name, Role, Username, Email, Note, Job_position FROM user_manager WHERE UserID = ?", (selected_id,))
            row = cursor.fetchone()
            if row:
                # Thiết lập kích thước dialog dựa trên vai trò người dùng
                dialog = customtkinter.CTkToplevel(root)
                dialog.title("Sửa Người Dùng")
                center_dialog(dialog)
                if row.Role.strip().lower() == 'admin':
                    dialog.geometry("400x500")
                else:
                    dialog.geometry("400x400")
                dialog.grab_set()  # Khóa dialog lại

                # Hiển thị thông tin và các trường chỉnh sửa
                customtkinter.CTkLabel(dialog, text=f"ID: {row.UserID}").place(relx=0.1, rely=0.05)
                customtkinter.CTkLabel(dialog, text="Tên:").place(relx=0.1, rely=0.12)
                name_entry = customtkinter.CTkEntry(dialog, width=310)
                name_entry.place(relx=0.1, rely=0.17)
                name_entry.insert(0, row.Name or "")

                customtkinter.CTkLabel(dialog, text="Vai trò:").place(relx=0.1, rely=0.27)
                role_label = customtkinter.CTkLabel(dialog, text=row.Role or "")
                role_label.place(relx=0.3, rely=0.27)

                # Lấy và hiển thị danh sách các vị trí công việc
                job_positions = get_job_positions()  # Lấy danh sách các vị trí từ hàm get_job_positions()
                customtkinter.CTkLabel(dialog, text="Vị trí:").place(relx=0.1, rely=0.33)
                position_combobox = customtkinter.CTkComboBox(dialog, values=job_positions, width=310)
                position_combobox.place(relx=0.1, rely=0.4)
                position_combobox.set(row.Job_position)  # Hiển thị vị trí hiện tại của người dùng

                # Chỉ hiển thị 'Tên Đăng Nhập' và 'Email' nếu role là 'admin'
                if row.Role.strip().lower() == 'admin':
                    customtkinter.CTkLabel(dialog, text="Tên Đăng Nhập:").place(relx=0.1, rely=0.5)
                    username_label = customtkinter.CTkLabel(dialog, text=row.Username or "")
                    username_label.place(relx=0.4, rely=0.5)

                    customtkinter.CTkLabel(dialog, text="Email:").place(relx=0.1, rely=0.55)
                    email_entry = customtkinter.CTkEntry(dialog, width=310)
                    email_entry.place(relx=0.1, rely=0.6)
                    email_entry.insert(0, row.Email if row.Email is not None else "")

                    note_y_position = 0.7
                else:
                    note_y_position = 0.5

                customtkinter.CTkLabel(dialog, text="Ghi Chú:").place(relx=0.1, rely=note_y_position)
                note_entry = customtkinter.CTkEntry(dialog, width=310)
                note_entry.place(relx=0.1, rely=note_y_position + 0.1)
                note_entry.insert(0, row.Note if row.Note is not None else "")

                # Label thông báo kết quả
                result_label = customtkinter.CTkLabel(dialog, text="")
                result_label.place(relx=0.1, rely=0.8)

                # Hàm lưu thông tin đã chỉnh sửa
                def save_changes():
                    new_name = name_entry.get()
                    new_email = email_entry.get() if row.Role.strip().lower() == 'admin' else ""  # Không lấy email nếu không phải admin
                    new_note = note_entry.get()
                    new_position = position_combobox.get()  # Lấy vị trí mới từ combobox

                    # Kiểm tra tính hợp lệ của email nếu role là admin
                    if row.Role.strip().lower() == 'admin' and not is_valid_email(new_email):
                        result_label.configure(text="Email không hợp lệ. Vui lòng kiểm tra lại.")
                        return  # Ngưng thực hiện nếu email không hợp lệ

                    # Nếu email hợp lệ, thực hiện lưu thay đổi
                    try:
                        cursor.execute(
                            "UPDATE user_manager SET Name = ?, Email = ?, Note = ?, Job_position = ? WHERE UserID = ?",
                            (new_name, new_email, new_note, new_position, row.UserID)
                        )
                        connection.commit()
                        result_label.configure(text="Thay đổi thông tin thành công!")
                        load_user_data()  # Tải lại dữ liệu người dùng
                        cursor.close()
                        connection.close()
                    except pyodbc.Error as ex:
                        # print("Error saving changes:", ex)
                        result_label.configure(text="Lỗi khi lưu thay đổi.")

                # Nút Lưu để cập nhật thông tin
                save_button = customtkinter.CTkButton(dialog, text="Lưu", command=save_changes, fg_color="green")
                save_button.place(relx=0.35, rely=0.9)



        except pyodbc.Error as ex:
            messagebox.showerror("Lỗi", "Lỗi khi sửa dữ liệu người dùng!")
            # print("Error loading user data for edit:", ex)
    else:
        messagebox.showerror("Lỗi", "Lỗi kết nối dữ liệu SQL!")
        # print("Server name or database name is missing.")

def delete_user():
    # Kiểm tra xem có chọn một bản ghi nào chưa
    selected_item = user_treeview.selection()
    if not selected_item:
        print("Vui lòng chọn một bản ghi để xóa.")
        return

    # Lấy ID của bản ghi đã chọn
    selected_id = user_treeview.item(selected_item, 'values')[0]

    # Tạo dialog xác nhận xóa
    confirm_dialog = customtkinter.CTkToplevel(root)
    confirm_dialog.title("Xác Nhận Xóa")
    center_dialog(confirm_dialog)
    confirm_dialog.geometry("300x150")
    
    message_label = customtkinter.CTkLabel(confirm_dialog, text="Bạn có muốn xóa bản ghi vừa chọn?")
    message_label.pack(pady=20)

    def confirm_delete():
        # Kết nối đến cơ sở dữ liệu và thực hiện xóa
        server_name, database_name = load_db_config()
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )
                cursor = connection.cursor()
                cursor.execute("DELETE FROM user_manager WHERE UserID = ?", (selected_id,))
                connection.commit()
                cursor.close()
                connection.close()
                messagebox.showinfo("Thành công","Xoá người dùng thành công!")
                # Cập nhật giao diện để xóa bản ghi
                load_user_data()  # Tải lại dữ liệu người dùng sau khi xóa
                confirm_dialog.destroy()  # Đóng dialog xác nhận
            except pyodbc.Error as ex:
                messagebox.showerror("Lỗi", "Lỗi khi xoá dữ liệu người dùng!")
                # print("Error deleting user data:", ex)

    def cancel_delete():
        confirm_dialog.destroy()  # Đóng dialog xác nhận

    # Nút Đồng Ý
# Nút Đồng Ý
    confirm_button = customtkinter.CTkButton(confirm_dialog, text="Đồng Ý", command=confirm_delete, fg_color="green", width=100)
    confirm_button.pack(side=tk.LEFT, padx=(40,5), pady=10)

    # Nút Huỷ
    cancel_button = customtkinter.CTkButton(confirm_dialog, text="Huỷ", command=cancel_delete, fg_color="red", width=100)
    cancel_button.pack(side=tk.RIGHT, padx=(5,40), pady=10)





root = tk.Tk()

root.geometry("1000x800")
root.title("Quản lý Admin")
root.resizable(False, False)    
center_dialog(root)


#doi icon cho frame chinh
icon_image = Image.open('icon_btn/setting.png')  # Thay đổi đường dẫn đến file icon của bạn
icon_image = icon_image.resize((20, 20), Image.LANCZOS)  # Kích thước icon
icon_photo = ImageTk.PhotoImage(icon_image)  # Tạo đối tượng PhotoImage
root.iconphoto(True, icon_photo)  # Đặt icon cho cửa sổ chính

#def tao icon cho nut
def load_icon(path, size=(30, 30)):  # Kích thước mặc định cho icon
    img = Image.open(path)
    img = img.resize(size, Image.LANCZOS)  # Thay đổi kích thước ảnh
    return ImageTk.PhotoImage(img)

def show_info():
    # Tạo dialog mới
    show_info_dialog = ctk.CTkToplevel()
    show_info_dialog.title("Thông Tin Người Dùng")
    show_info_dialog.geometry("400x300")
    center_dialog(show_info_dialog)
    show_info_dialog.grab_set()

    # UserID
    ctk.CTkLabel(show_info_dialog, text="UserID:").place(relx=0.1, rely=0.05)
    user_id_entry = ctk.CTkEntry(show_info_dialog, width=200)
    user_id_entry.place(relx=0.25, rely=0.05)
    
    # Role
    ctk.CTkLabel(show_info_dialog, text="Vai Trò:").place(relx=0.1, rely=0.25)
    role_entry = ctk.CTkEntry(show_info_dialog, width=200, state='readonly')
    role_entry.place(relx=0.25, rely=0.25)

    # Name
    ctk.CTkLabel(show_info_dialog, text="Tên:").place(relx=0.1, rely=0.45)
    name_entry = ctk.CTkEntry(show_info_dialog, width=200, state='readonly')
    name_entry.place(relx=0.25, rely=0.45)

    # Email    
    ctk.CTkLabel(show_info_dialog, text="Email:").place(relx=0.1, rely=0.65)
    email_entry = ctk.CTkEntry(show_info_dialog, width=200, state='readonly')
    email_entry.place(relx=0.25, rely=0.65)
    
    user_id = get_userid()  # Gọi hàm get_userid() để lấy UserID của người dùng

    # Điền UserID vào entry mà không cần người dùng nhập
    user_id_entry.insert(0, str(user_id))  # Chèn UserID vào ô nhập liệu

    # Lấy thông tin người dùng
    def get_user_info():
        # user_id = user_id_entry.get()
        server_name, database_name = load_db_config()
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )
                cursor = connection.cursor()
                cursor.execute("SELECT Name, Role, Email FROM user_manager WHERE UserID = ?", user_id)
                result = cursor.fetchone()
                if result:
                    user_id_entry.configure(state='readonly')
                    name_entry.configure(state='normal')
                    email_entry.configure(state='normal')
                    name_entry.delete(0, 'end')
                    email_entry.delete(0, 'end')
                    name_entry.insert(0, result[0])
                    role_entry.configure(state='normal')
                    role_entry.delete(0, 'end')
                    role_entry.insert(0, result[1])
                    role_entry.configure(state='readonly')
                    email_value = result[2] if result[2] is not None else ""  # Nếu email là None thì gán là chuỗi rỗng
                    email_entry.insert(0, email_value)  # Chèn email
                    name_entry.configure(state='readonly')
                    email_entry.configure(state='readonly')
                else:
                    print("Không tìm thấy người dùng.")
                cursor.close()
                connection.close()
            except pyodbc.Error as ex:
                messagebox.showerror("Lỗi", "Lỗi tải dữ liệu người dùng!")
                # print("Error loading user info:", ex)

    get_user_info()

    # Nút Thay Đổi Thông Tin
    change_info_button = ctk.CTkButton(show_info_dialog, text="Thay Đổi Thông Tin", 
                                        command=lambda: edit_user_info(user_id_entry.get()))
    change_info_button.place(relx=0.52, rely=0.8)

    # Nút Đổi Mật Khẩu
    change_password_button = ctk.CTkButton(show_info_dialog, text="Đổi Mật Khẩu", 
                                            command=lambda: change_password_dialog(user_id_entry.get(),name_entry.get()))
    change_password_button.place(relx=0.12, rely=0.8)

def edit_user_info(user_id):
    # Tạo dialog mới
    edit_user_dialog = ctk.CTkToplevel()
    edit_user_dialog.title("Chỉnh Sửa Thông Tin Người Dùng")
    edit_user_dialog.geometry("400x350")
    center_dialog(edit_user_dialog)
    edit_user_dialog.grab_set()

    # UserID
    ctk.CTkLabel(edit_user_dialog, text="UserID:").place(relx=0.1, rely=0.05)
    user_id_entry = ctk.CTkEntry(edit_user_dialog, width=200)  # Chỉ đọc
    user_id_entry.place(relx=0.25, rely=0.05)
    user_id_entry.insert(0, user_id)
    user_id_entry.configure(state='disabled')
    # print(f"UserID: {user_id}")


    # Role
    ctk.CTkLabel(edit_user_dialog, text="Vai Trò:").place(relx=0.1, rely=0.25)
    role_entry = ctk.CTkEntry(edit_user_dialog, width=200, state='readonly')
    role_entry.place(relx=0.25, rely=0.25)

    # Name
    ctk.CTkLabel(edit_user_dialog, text="Tên:").place(relx=0.1, rely=0.45)
    name_entry = ctk.CTkEntry(edit_user_dialog, width=200)  # Có thể chỉnh sửa
    name_entry.place(relx=0.25, rely=0.45)

    # Email    
    ctk.CTkLabel(edit_user_dialog, text="Email:").place(relx=0.1, rely=0.65)
    email_entry = ctk.CTkEntry(edit_user_dialog, width=200)  # Có thể chỉnh sửa
    email_entry.place(relx=0.25, rely=0.65)

    # Lấy thông tin người dùng
    def get_user_info():
        server_name, database_name = load_db_config()
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )
                cursor = connection.cursor()
                cursor.execute("SELECT Name, Role, Email FROM user_manager WHERE UserID = ?", user_id)
                result = cursor.fetchone()
                if result:
                    role_entry.configure(state='normal')
                    role_entry.delete(0, 'end')
                    role_entry.insert(0, result[1])
                    role_entry.configure(state='readonly')

                    name_entry.delete(0, 'end')
                    name_entry.insert(0, result[0])  # Cập nhật tên
                    email_entry.delete(0, 'end')
                    email_value = result[2] if result[2] is not None else ""  # Nếu email là None thì gán là chuỗi rỗng
                    email_entry.insert(0, email_value)  # Chèn email  # Cập nhật email
                else:
                    print("Không tìm thấy người dùng.")
                cursor.close()
                connection.close()
            except pyodbc.Error as ex:
                messagebox.showerror("Lỗi", "Lỗi tải dữ liệu người dùng!")
                # print("Error loading user info:", ex)

    # Gọi hàm lấy thông tin khi khởi tạo dialog
    get_user_info()

    # Nút Lưu Thay Đổi
    save_changes_button = ctk.CTkButton(edit_user_dialog, text="Lưu Thay Đổi", 
                                         command=lambda: change_user_info(user_id, 
                                                                          name_entry.get(), 
                                                                          email_entry.get()))
    save_changes_button.place(relx=0.52, rely=0.8)

    # Nút Hủy
    cancel_button = ctk.CTkButton(edit_user_dialog, text="Hủy", command=edit_user_dialog.destroy)
    cancel_button.place(relx=0.12, rely=0.8)

def change_user_info(user_id, name, email):
    server_name, database_name = load_db_config()
    if server_name and database_name:
        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )
            cursor = connection.cursor()
            # Cập nhật thông tin người dùng
            cursor.execute("""
                UPDATE user_manager
                SET Name = ?, Email = ?
                WHERE UserID = ?
            """, (name, email, user_id))
            connection.commit()
            print("Cập nhật thông tin thành công.")
            cursor.close()
            connection.close()
        except pyodbc.Error as ex:
            messagebox.showerror("Lỗi", "Cập nhật dữ liệu người dùng thất bại"+ex)
            # print("Error updating user info:", ex)


def change_password_dialog(user_id, name):
    # Tạo dialog mới
    password_dialog = ctk.CTkToplevel()
    password_dialog.title("Đổi Mật Khẩu")
    password_dialog.geometry("500x400")
    center_dialog(password_dialog)
    password_dialog.grab_set()

    # Hiển thị UserID và Name
    ctk.CTkLabel(password_dialog, text="UserID:").place(relx=0.1, rely=0.05)
    user_id_entry = ctk.CTkEntry(password_dialog, width=200)
    user_id_entry.place(relx=0.4, rely=0.05)
    user_id_entry.insert(0, user_id)
    # print(f"UserID: {user_id}")

    ctk.CTkLabel(password_dialog, text="Name:").place(relx=0.1, rely=0.2)
    name_entry = ctk.CTkEntry(password_dialog, width=200)
    name_entry.place(relx=0.4, rely=0.2)
    name_entry.insert(0, name)
    # print(f"Name: {name}")

    # Trường mật khẩu cũ
    ctk.CTkLabel(password_dialog, text="Mật Khẩu Cũ:").place(relx=0.1, rely=0.35)
    old_password_entry = ctk.CTkEntry(password_dialog, width=200, show="*")
    old_password_entry.place(relx=0.4, rely=0.35)
    old_password_check_label = ctk.CTkLabel(password_dialog, text="")  # Label trống để hiển thị tích xanh
    old_password_check_label.place(relx=0.8, rely=0.35)
    
    # Trường mật khẩu mới
    ctk.CTkLabel(password_dialog, text="Mật Khẩu Mới:").place(relx=0.1, rely=0.5)
    new_password_entry = ctk.CTkEntry(password_dialog, width=200, show="*")
    new_password_entry.place(relx=0.4, rely=0.5)
    
    def toggle_password_visibility():
        if new_password_entry.cget('show') == '*':
            new_password_entry.configure(show='')
            toggle_button.configure(text='Hiện')
        else:
            new_password_entry.configure(show='*')
            toggle_button.configure(text='Ẩn đi')

    toggle_button = ctk.CTkButton(password_dialog, text='Hiện', command=toggle_password_visibility, width=30, height=28)
    toggle_button.place(relx=0.83, rely=0.5)

    # Trường lặp lại mật khẩu mới
    ctk.CTkLabel(password_dialog, text="Lặp Lại Mật Khẩu Mới:").place(relx=0.1, rely=0.65)
    repeat_password_entry = ctk.CTkEntry(password_dialog, width=200, show="*")
    repeat_password_entry.place(relx=0.4, rely=0.65)

    # Hàm kiểm tra mật khẩu cũ
    def check_old_password():
        acpt_icon = CTkImage(Image.open("icon/acpt.png"), size=(30, 30))
        deny_icon = CTkImage(Image.open("icon/deny.png"), size=(25, 25))
        old_password = old_password_entry.get()
        # Kiểm tra mật khẩu cũ trong cơ sở dữ liệu
        if verify_old_password(user_id, old_password):
            old_password_check_label.configure(image=acpt_icon)
            old_password_check_label.place(relx=0.85,rely=0.35)# Tích xanh nếu đúng
        else:
            old_password_check_label.configure(image=deny_icon)  # Dấu X nếu sai
            old_password_check_label.place(relx=0.85,rely=0.35)
    # Gán sự kiện check mật khẩu cho trường mật khẩu cũ
    old_password_entry.bind("<FocusOut>", lambda event: check_old_password())

    # Hàm lưu thay đổi mật khẩu
    def save_password_change():
        if not verify_old_password(user_id, old_password_entry.get()):
            messagebox.showerror("Lỗi", "Mật khẩu cũ không đúng!")
            # print("Mật khẩu cũ không đúng.")
            return
        if new_password_entry.get() != repeat_password_entry.get():
            messagebox.showerror("Lỗi", "Mật khẩu mới không khớp!")
            # print("Mật khẩu mới không khớp.")
            return
        # Thay đổi mật khẩu trong cơ sở dữ liệu
        update_password(user_id, new_password_entry.get())
        messagebox.showinfo("Thông báo", "Đổi mật khẩu thành công!")
        # print("pw updated")
        password_dialog.destroy()

    # Nút lưu thay đổi
    save_button = ctk.CTkButton(password_dialog, text="Lưu Thay Đổi", command=save_password_change)
    save_button.place(relx=0.55, rely=0.8)

    # Nút hủy
    cancel_button = ctk.CTkButton(password_dialog, text="Hủy", command=password_dialog.destroy)
    cancel_button.place(relx=0.2, rely=0.8)

# Hàm kiểm tra mật khẩu cũ từ cơ sở dữ liệu

def verify_old_password(user_id, old_password):
    # Kết nối đến cơ sở dữ liệu
    server_name, database_name = load_db_config()  # Hàm lấy thông tin server và database
    if not (server_name and database_name):
        messagebox.showerror("Lỗi", "Thiếu thông tin cấu hình Servername hoặc DB!")
            # print("Error updating user info:", ex)
        # print("Thiếu thông tin cấu hình server hoặc database.")
        return False

    try:
        connection = pyodbc.connect(
            'DRIVER={SQL Server};' +
            f'Server={server_name};' +
            f'Database={database_name};' +
            'Trusted_Connection=True'
        )
        cursor = connection.cursor()

        # Lấy mật khẩu hiện tại của user từ cơ sở dữ liệu
        query = "SELECT Password FROM user_manager WHERE UserID = ?"
        cursor.execute(query, user_id)
        result = cursor.fetchone()

        # Nếu user tồn tại và mật khẩu khớp
        if result and bcrypt.checkpw(old_password.encode('utf-8'), result[0].encode('utf-8')):
            return True
        else:
            return False

    except pyodbc.Error as ex:
        messagebox.showerror("Lỗi", "Lỗi khi kiểm tra mật khẩu!"+ex)
            # print("Error updating user info:", ex)
        # print("Lỗi khi kiểm tra mật khẩu:", ex)
        return False
    finally:
        cursor.close()
        connection.close()


# Hàm cập nhật mật khẩu mới trong cơ sở dữ liệu

def update_password(user_id, new_password):
    # Kết nối đến cơ sở dữ liệu
    server_name, database_name = load_db_config()  # Hàm lấy thông tin server và database
    if not (server_name and database_name):
        messagebox.showerror("Lỗi", "Thiếu thông tin cấu hình Servername hoặc DB!")
        # print("Thiếu thông tin cấu hình server hoặc database.")
        return False

    try:
        hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
        hashed_password_str = hashed_password.decode('utf-8')
        
        connection = pyodbc.connect(
            'DRIVER={SQL Server};' +
            f'Server={server_name};' +
            f'Database={database_name};' +
            'Trusted_Connection=True'
        )
        cursor = connection.cursor()

        # Cập nhật mật khẩu mới cho user dựa trên user_id
        query = "UPDATE user_manager SET Password = ? WHERE UserID = ?"
        cursor.execute(query, hashed_password_str, user_id)
        
        # Xác nhận thay đổi
        connection.commit()

        print("Mật khẩu đã được cập nhật thành công.")
        return True

    except pyodbc.Error as ex:
        print("Lỗi khi cập nhật mật khẩu:", ex)
        return False
    finally:
        cursor.close()
        connection.close()

















def logout():
    print("Đăng xuất")
    
    # root.destroy()
    create_login_dialog()
    # subprocess.Popen(["python", "admin_login.py"])
    
    # Hàm xử lý cho "Đăng xuất"
    # 
    


# Tạo icon cho các nút
user_manager_icon = load_icon('icon_btn/userbtn.png')  # Thay đổi đường dẫn đến file icon của bạn
att_manager_icon = load_icon('icon_btn/user2btn.png')  # Thay đổi đường dẫn đến file icon của bạn
catalog_icon = load_icon('icon_btn/catabtn.png')  # Thay đổi đường dẫn đến file icon của bạn

#tao icon cho cac nut chuc nang nho khac
load_btn = load_icon('icon_btn/load_data.png')
add_btn = load_icon('icon_btn/add_btn.png')
add_face_btn=load_icon('icon/face_icon.png')
edit_btn = load_icon('icon_btn/edit_btn.png')
delete_btn = load_icon('icon_btn/remove_btn.png')   
search_btn = load_icon('icon_btn/search_btn.png')
filter_btn = load_icon('icon_btn/filter_btn.png')
export_btn = load_icon('icon_btn/export_btn.png')
# Header
header_frame = tk.Frame(root, width=250,bd=20,bg='lightblue')
header_frame.pack(fill='x')

separator = tk.Frame(root, height=2, bd=1, bg='black')  # Đường kẻ ngang với màu đen
separator.pack(fill='x')

# Thời gian và ngày hiện tại
time_label = tk.Label(header_frame, font=("Arial", 16))
time_label.pack(side="left", padx=50)
date_label = tk.Label(header_frame, font=("Arial", 16))
date_label.pack(side="left", padx=50)


# greeting_label = tk.Label(header_frame, text="Xin chào, ADMIN", font=("Arial", 18))
# greeting_label.pack(side="right", padx=50)

admin_button = tk.Menubutton(header_frame, text="ADMIN", font=("Arial", 18),padx=10,pady=10,width=8,relief="raised", bd=2)
admin_button.menu = tk.Menu(admin_button, tearoff=0)
admin_button['menu'] = admin_button.menu
admin_button.menu.add_command(label="Xem thông tin", command=show_info)
admin_button.menu.add_separator()
admin_button.menu.add_command(label="Đăng xuất", command=logout)
admin_button.pack(side="right", padx=50)

db_check_button = tk.Button(header_frame, text="Kiểm tra kết nối DB", command=check_db_connection, bg="green", fg="white", padx=10, pady=5)
db_check_button.pack(side="right", padx=20)

# Menu bên trái
menu_frame = tk.Frame(root,width=170, bd=20)
menu_frame.pack(side="left", fill='y')

user_management_button = tk.Button(menu_frame, text="Quản lý người dùng", command=show_user_management, height=130, width=170,image=user_manager_icon, compound='left',padx=15)
user_management_button.pack(fill='both',padx=10, pady=(40, 40))

attendance_management_button = tk.Button(menu_frame, text="Quản lý dữ liệu điểm danh", command=show_attendance_management,height=130, width=170,image=att_manager_icon, compound='left',padx=15)
attendance_management_button.pack(fill='both',padx=10, pady=(30, 40))

catalog_button = tk.Button(menu_frame, text="Quản lý vai trò người dùng", command=show_catalog_management,height=130, width=170,image=catalog_icon, compound='left',padx=15)
catalog_button.pack(fill='both',padx=10, pady=(30, 40))

# Khung quản lý người dùng
user_management_frame = tk.Frame(root, width=1000, height=850)
user_management_frame.pack(fill='both', expand=True)

# Nút chức năng quản lý người dùng
user_function_frame = tk.Frame(user_management_frame)
user_function_frame.pack(side="top", anchor="center", padx=10, pady=10)


add_button = tk.Button(user_function_frame, text="Tải DL",image=load_btn, compound='left',padx=8,command=load_user_data)
add_button.pack(side="left", padx=(0,5))

add_button = tk.Button(user_function_frame, text="Thêm",image=add_btn, compound='left',padx=8,command=add_user)
add_button.pack(side="left", padx=10)

add_button = tk.Button(user_function_frame, text="Thêm khuôn mặt",image=add_face_btn, compound='left',padx=8,command=add_face_user)
add_button.pack(side="left", padx=10)

edit_button = tk.Button(user_function_frame, text="Sửa",image=edit_btn, compound='left',padx=8,command=edit_user)
edit_button.pack(side="left", padx=10)

delete_button = tk.Button(user_function_frame, text="Xóa",image=delete_btn, compound='left',padx=8,command=delete_user)
delete_button.pack(side="left", padx=10)

search_button = tk.Button(user_function_frame, text="Tìm kiếm",image=search_btn, compound='left',padx=10,command=dialog_search_data)
search_button.pack(side="left", padx=(10,0))

# Cột dữ liệu
columns = ('UserID', 'Tên', 'Vai trò','Vị trí','Tên đăng nhập','Email', 'Ghi chú','Check')
user_treeview = ttk.Treeview(user_management_frame, columns=columns, show='headings')

for col in columns:
    user_treeview.heading(col, text=col)
user_treeview.column('UserID', width=20,anchor='center')  # Độ rộng cho cột ID
user_treeview.column('Tên', width=80)  # Độ rộng cho cột Tên
user_treeview.column('Vai trò', width=30) 
user_treeview.column('Vị trí', width=50) 
user_treeview.column('Tên đăng nhập', width=50)
user_treeview.column('Email', width=100)
user_treeview.column('Ghi chú', width=100)
user_treeview.column('Check', width=40,anchor='center') # Độ rộng cho cột Ghi chú
user_treeview.pack(fill='both', expand=True)




#QUẢN LÝ DỮ LIỆU ĐIỂM DANH

#function of attention_manager
def load_attention_data():
    # Kết nối đến cơ sở dữ liệu
    selected_month = month_combo.get()
    server_name, database_name = load_db_config()  # Đảm bảo rằng hàm này trả về thông tin kết nối chính xác
    if server_name and database_name:
        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )
            cursor = connection.cursor()
            query=f"""SELECT UserID, Name,job_position,AttentionDate, CheckInTime, CheckOutTime, Status FROM attention_manager WHERE MONTH(AttentionDate) = {selected_month}"""
            cursor.execute(query)
            rows = cursor.fetchall()
            cursor.close()
            connection.close()

            # Xóa tất cả dữ liệu trong Treeview trước khi thêm mới
            for item in attendance_treeview.get_children():
                attendance_treeview.delete(item)

            # Thêm dữ liệu vào Treeview
            for row in rows:
                check_in_time_str = row[4].split('.')[0] if row[4] else None  # Lấy phần trước dấu chấm
                check_out_time_str = row[5].split('.')[0] if row[5] else None  # Lấy phần trước dấu chấm
                
                check_in_time = (datetime.strptime(check_in_time_str, "%H:%M:%S").strftime("%H:%M") if check_in_time_str else None)
                check_out_time = (datetime.strptime(check_out_time_str, "%H:%M:%S").strftime("%H:%M") if check_out_time_str else None)
                attendance_treeview.insert("", "end", values=(row[0], row[1], row[2], row[3], check_in_time,check_out_time,row[6]))

        except pyodbc.Error as ex:
            messagebox.showerror("Lỗi", "Lỗi tải dữ liệu!")+ex
            # print("Error loading data:", ex)
    else:
        messagebox.showerror("Lỗi", "Thiếu thông tin cấu hình Servername hoặc DB!")
        # print("Server name or database name is missing.")


# def dialog_search_data2():
#     search_dialog = tk.Toplevel(root)  # Tạo cửa sổ mới
#     search_dialog.title("Tìm Kiếm")
#     search_dialog.geometry("400x250")
#     search_dialog.resizable(False, False)
#     search_dialog.grab_set()

#     center_dialog(search_dialog)
#     # Các trường nhập liệu
#     labels = ['UserID', 'Tên', 'Vai trò', 'Tên đăng nhập', 'Email', 'Ghi chú']
#     entries = []

#     for label in labels:
#         row = tk.Frame(search_dialog)
#         lab = tk.Label(row, text=label, width=15, anchor='w')  # Căn trái
#         ent = tk.Entry(row, width=30)
#         lab.pack(side=tk.LEFT, padx=(10, 5), pady=(2, 0))  # Khoảng cách 10px bên trái, 5px bên phải
#         ent.pack(side=tk.RIGHT, padx=(0, 10), pady=(2, 0))  # Khoảng cách 10px bên phải
#         row.pack(pady=5)
#         entries.append(ent)

#     # Nút tìm kiếm
#     search_button = tk.Button(search_dialog, text="Tìm Kiếm", command=lambda: search_data2(entries, search_dialog), width=20)
#     search_button.pack(pady=10)

#     def user_data_search():
#         # Kết nối đến cơ sở dữ liệu
#         server_name, database_name = load_db_config()
#         if server_name and database_name:
#             try:
#                 connection = pyodbc.connect(
#                     'DRIVER={SQL Server};' +
#                     f'Server={server_name};' +
#                     f'Database={database_name};' +
#                     'Trusted_Connection=True'
#                 )

#                 cursor = connection.cursor()
#                 cursor.execute("SELECT UserID, Name, Role, Username, Email, Note FROM user_manager")  # Lấy 6 cột
#                 rows = cursor.fetchall()
#                 cursor.close()
#                 connection.close()

#                 return rows  # Trả về dữ liệu để sử dụng trong các hàm khác
#             except pyodbc.Error as ex:
#                 messagebox.showerror("Lỗi", "Lỗi tải dữ liệu người dùng!")
#                 # print("Error loading user data:", ex)
#                 return []  # Trả về danh sách rỗng nếu có lỗi
#         else:
#             messagebox.showerror("Lỗi", "Thiếu thông tin cấu hình Servername hoặc DB!")
#             # print("Server name or database name is missing.")
#             return []  # Trả về danh sách rỗng nếu không có thông tin kết nối

#     def search_data2(entries, search_dialog):
#         search_terms = [entry.get().lower() for entry in entries]  
#         for item in user_treeview.get_children():
#             user_treeview.delete(item)

#         all_data = user_data_search()  # Sử dụng load_user_data() để lấy dữ liệu từ SQL
#         found_records = False  # Biến đánh dấu có tìm thấy bản ghi hay không

#         for record in all_data:
#             conditions = [
#                 (not search_terms[0] or str(record[0]) == search_terms[0]),  # ID
#                 (not search_terms[1] or search_terms[1] in record[1].lower()),  # Tên
#                 (not search_terms[2] or search_terms[2] in record[2].lower()),  # Vai trò
#                 (not search_terms[3] or search_terms[3] in record[3].lower()),  # Tên đăng nhập
#                 (not search_terms[4] or search_terms[4] in record[4].lower()),  # Email
#                 (not search_terms[5] or search_terms[5] in record[5].lower()),  # Ghi chú
#             ]
            
#             if all(conditions):  # Nếu tất cả các điều kiện thỏa mãn
#                 user_treeview.insert("", "end", values=(record[0], record[1], record[2], record[3], record[4], record[5]))
#                 found_records = True  # Đánh dấu là đã tìm thấy bản ghi

#         if not found_records:  # Nếu không tìm thấy bản ghi nào
#             show_no_record_found_dialog()
#         else:
#             search_dialog.destroy()  # Chỉ đóng dialog tìm kiếm nếu có bản ghi được tìm thấy

#     def show_no_record_found_dialog():
#         no_record_dialog = tk.Toplevel(root)
        
#         no_record_dialog.title("Thông báo")
#         center_dialog(no_record_dialog)
#         no_record_dialog.geometry("300x150")  # Thiết lập kích thước cố định trước
#         # Căn giữa sau khi thiết lập kích thước
#         no_record_dialog.grab_set()  # Đảm bảo dialog này nhận tất cả sự chú ý
        
#         message = tk.Label(no_record_dialog, text="Không tìm thấy bản ghi phù hợp!", padx=10, pady=10)
#         message.pack()

#         ok_button = tk.Button(no_record_dialog, text="OK", command=no_record_dialog.destroy,width=15, height=2,bg="green")
#         ok_button.pack(pady=10, padx=10)

#         no_record_dialog.focus_set()  # Đặt focus cho dialog thông báo
#         center_dialog(no_record_dialog)




def open_filter_dialog():
    # Tạo dialog mới
    filter_dialog = ctk.CTkToplevel()
    filter_dialog.title("Lọc Dữ Liệu")
    filter_dialog.geometry("500x350")  # Tăng kích thước dialog để có thêm không gian
    center_dialog(filter_dialog)  # Hàm này cần được định nghĩa trước
    filter_dialog.grab_set()

    # UserID
    ctk.CTkLabel(filter_dialog, text="UserID:").grid(row=0, column=0, padx=10, pady=(10,5))
    user_id_entry = ctk.CTkEntry(filter_dialog, width=230)
    user_id_entry.grid(row=0, column=1, padx=40, pady=5)

    # Name
    ctk.CTkLabel(filter_dialog, text="Tên:").grid(row=1, column=0, padx=10, pady=(10,5))
    name_entry = ctk.CTkEntry(filter_dialog, width=230)
    name_entry.grid(row=1, column=1, padx=30, pady=5)
    
    # AttentionDate
    ctk.CTkLabel(filter_dialog, text="Ngày:").grid(row=2, column=0, padx=10, pady=(10,5))
    date_entry = ctk.CTkEntry(filter_dialog, width=230)
    date_entry.grid(row=2, column=1, padx=20, pady=5)

    # Month (ComboBox for month)
    ctk.CTkLabel(filter_dialog, text="Tháng:").grid(row=3, column=0, padx=10, pady=(10,5))
    month_combo = ctk.CTkComboBox(filter_dialog, values=[f"{i+1:02d}" for i in range(12)], width=230)
    month_combo.set(f"{datetime.today().month:02d}")  # Set current month by default
    month_combo.grid(row=3, column=1, padx=20, pady=5)

    # Status (ComboBox for status)
    ctk.CTkLabel(filter_dialog, text="Trạng Thái:").grid(row=4, column=0, padx=10, pady=(10,5))
    status_combobox = ctk.CTkComboBox(filter_dialog, values=["Tất cả", "Muon", "Da diem danh"], width=230)
    status_combobox.set("Tất cả")  # Set default value to "Tất cả"
    status_combobox.grid(row=4, column=1, padx=20, pady=5)

    # Position (ComboBox for position from job_position table)
    ctk.CTkLabel(filter_dialog, text="Vị trí:").grid(row=5, column=0, padx=10, pady=(10,5))
    position_combobox = ctk.CTkComboBox(filter_dialog, width=230)
    position_combobox.grid(row=5, column=1, padx=20, pady=5)

    # Lấy danh sách vị trí từ bảng job_position
    def load_positions():
        server_name, database_name = load_db_config()  # Đảm bảo hàm này trả về thông tin kết nối chính xác
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )
                cursor = connection.cursor()
                
                # Lấy tất cả các vị trí từ bảng job_position
                cursor.execute("SELECT DISTINCT job_position FROM job_position")
                rows = cursor.fetchall()
                cursor.close()
                connection.close()

                # Lấy tất cả vị trí và điền vào ComboBox
                positions = [row[0] for row in rows]
                position_combobox.configure(values=positions)
                position_combobox.set("Tất cả")  # Đặt mặc định là "Tất cả" nếu không có vị trí nào được chọn

            except pyodbc.Error as ex:
                messagebox.showerror("Lỗi", f"Lỗi kết nối cơ sở dữ liệu: {ex}")
        else:
            messagebox.showerror("Lỗi", "Không thể tải thông tin kết nối cơ sở dữ liệu!")

    # Gọi hàm tải vị trí khi mở dialog
    load_positions()

    # Nút Lọc
    def filter_data():
        user_id = user_id_entry.get()
        name = name_entry.get()
        attention_date = date_entry.get()
        status = status_combobox.get()
        position = position_combobox.get()  # Lấy giá trị vị trí từ ComboBox

        # Lấy năm hiện tại
        current_year = datetime.today().year  # Lấy năm hiện tại
        selected_month = month_combo.get()

        try:
            # Kiểm tra ngày theo tháng đã chọn
            if attention_date:
                day = int(attention_date)  # Lấy ngày nhập từ user
                month = int(selected_month)
                if not check_valid_date(day, month, current_year):
                    messagebox.showerror("Lỗi", f"Ngày {attention_date} không hợp lệ trong tháng {selected_month}!")
                    return
                # Nếu ngày hợp lệ, chuẩn hóa ngày
                normalized_date = f"{current_year}-{selected_month}-{attention_date.zfill(2)}"
            else:
                # normalized_date = None  # Nếu không nhập ngày thì để trống
                normalized_date = f"{current_year}-{selected_month.zfill(2)}%"
    
            # Lọc dữ liệu
            apply_filter(user_id, name, normalized_date, status, position)

            filter_dialog.destroy()  # Đóng dialog sau khi lọc

        except ValueError:
            messagebox.showerror("Lỗi", "Ngày không hợp lệ, vui lòng nhập đúng định dạng số!")

    filter_button = ctk.CTkButton(filter_dialog, text="Lọc", command=filter_data)
    filter_button.place(relx=0.5, rely=0.85, anchor="center")

    def check_valid_date(day, month, year):
        try:
            # Thử tạo một đối tượng datetime với ngày, tháng, năm
            datetime(year, month, day)
            return True
        except ValueError:
            return False

    def apply_filter(user_id, name, attention_date, status, position):
        for item in attendance_treeview.get_children():
            attendance_treeview.delete(item)  # Xóa tất cả dữ liệu hiện tại

        # Kết nối đến cơ sở dữ liệu
        server_name, database_name = load_db_config()  # Đảm bảo hàm này trả về thông tin kết nối chính xác
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )
                cursor = connection.cursor()
                
                # Tạo câu lệnh SQL lọc
                query = "SELECT UserID, Name, job_position, AttentionDate, CheckInTime, CheckOutTime, Status FROM attention_manager WHERE 1=1"
                params = []

                if user_id:
                    query += " AND UserID = ?"
                    params.append(user_id)
                if name:
                    query += " AND Name LIKE ?"
                    params.append(f'%{name}%')
                if attention_date:
                    query += " AND AttentionDate LIKE ?"
                    params.append(attention_date)
                if status and status != "Tất cả":
                    query += " AND Status = ?"
                    params.append(status)
                if position and position != "Tất cả":
                    query += " AND job_position LIKE ?"
                    params.append(f'%{position}%')  # Tìm kiếm theo vị trí

                cursor.execute(query, params)
                rows = cursor.fetchall()
                cursor.close()
                connection.close()

                # Đếm số bản ghi tìm được
                if rows:
                    messagebox.showinfo("Kết quả", f"Số bản ghi tìm được: {len(rows)}")
                else:
                    messagebox.showinfo("Kết quả", "Không tìm thấy bản ghi nào!")

                # Thêm dữ liệu vào Treeview
                for row in rows:
                    check_in_time_str = row[4].split('.')[0] if row[4] else None  # Lấy phần trước dấu chấm
                    check_out_time_str = row[5].split('.')[0] if row[5] else None  # Lấy phần trước dấu chấm
                    
                    check_in_time = (datetime.strptime(check_in_time_str, "%H:%M:%S").strftime("%H:%M") if check_in_time_str else None)
                    check_out_time = (datetime.strptime(check_out_time_str, "%H:%M:%S").strftime("%H:%M") if check_out_time_str else None)
                    
                    attendance_treeview.insert("", "end", values=(row[0], row[1], row[2], row[3], check_in_time, check_out_time, row[6]))

            except pyodbc.Error as ex:
                messagebox.showerror("Lỗi", f"Lỗi kết nối cơ sở dữ liệu: {ex}")
                

def export_to_excel():
    # Lấy tháng từ month_combo
    month = month_combo.get()
    records = []
    for item in attendance_treeview.get_children():
        record = attendance_treeview.item(item)['values']
        records.append(record)

    if not records:  # Kiểm tra nếu Treeview trống
        messagebox.showwarning("Cảnh báo", "Bạn cần tải dữ liệu cần xuất Excel trước!")
        return  # Dừng hàm nếu không có dữ liệu

    # Mở hộp thoại để chọn đường dẫn lưu file với tên mặc định là 'Dulieudiemdanh.xlsx'
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                               title="Lưu file Excel", 
                                               initialfile=f"Dulieudiemdanh_thang{month}.xlsx")

    if file_path:  # Kiểm tra nếu người dùng đã chọn đường dẫn
        try:
            # Tạo một workbook và sheet mới
            wb = openpyxl.Workbook()
            ws = wb.active

            # Lấy năm hiện tại
            current_date = datetime.now()
            current_year = current_date.year

            # Lấy ngày cuối tháng của tháng được chọn trong month_combo
            _, last_day = calendar.monthrange(current_year, int(month))
            
            # Hàng 1: Tiêu đề lớn
            title = f"Dữ liệu điểm danh tháng {month}"
            ws.merge_cells('A1:G1')  # Hợp nhất các ô từ A1 đến G1
            ws['A1'] = title  # Đặt tiêu đề vào ô A1
            title_cell = ws['A1']
            title_cell.font = Font(size=16, bold=True)  # Kích thước và đậm
            title_cell.alignment = Alignment(horizontal='center')  # Căn giữa

            # Hàng 2: Căn giữa từ A2 đến G2, với nội dung ngày tháng từ month_combo
            date_range = f"Từ 01/{int(month):02d}/{current_year} - {last_day:02d}/{int(month):02d}/{current_year}"
            ws.merge_cells('A2:G2')  # Hợp nhất các ô từ A2 đến G2
            ws['A2'] = date_range  # Đặt nội dung
            ws['A2'].alignment = Alignment(horizontal='center')  # Căn giữa

            # Hàng 3: Gộp từ A3 đến G3, căn phải, in nghiêng
            ws.merge_cells('A3:G3')  # Hợp nhất các ô từ A3 đến G3
            footer_text = "Dữ liệu được lấy từ phần mềm điểm danh khuôn mặt Face_Recognition v1.0.1"
            ws['A3'] = footer_text  # Đặt nội dung
            ws['A3'].alignment = Alignment(horizontal='right')  # Căn phải
            ws['A3'].font = Font(italic=True)  # In nghiêng

            # Thêm tiêu đề cột ở hàng thứ 5 (sau khi hàng 3 đã gộp)
            ws.append([])  # Thêm hàng trống thứ 2
            ws.append(["UserID", "Tên", "Vị trí", "Ngày", "Giờ vào", "Giờ ra", "Trạng thái"])

            # Tô màu xen kẽ các dòng có dữ liệu từ hàng 5 trở đi
            fill1 = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Màu cho dòng chẵn
            fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Màu cho dòng lẻ
            border = Border(left=Side(border_style="thin"),
                            right=Side(border_style="thin"),
                            top=Side(border_style="thin"),
                            bottom=Side(border_style="thin"))  # Viền mỏng cho các ô

            # Duyệt qua dữ liệu và thêm vào worksheet
            for i, record in enumerate(records):
                # Chọn màu nền cho dòng chẵn và lẻ, bắt đầu từ hàng 5 (hàng đầu tiên có dữ liệu)
                fill = fill1 if i % 2 == 0 else fill2
                # Thêm dữ liệu và tô màu nền cho dòng
                row_idx = i + 6  # Dòng bắt đầu từ hàng 5
                ws.append(record)
                for j in range(7):  # 7 cột trong mỗi dòng
                    cell = ws.cell(row=row_idx, column=j+1)
                    cell.fill = fill  # Áp dụng màu nền cho từng ô
                    cell.border = border  # Áp dụng viền cho từng ô
                    if j == 0 or j == 4 or j == 5:  # Căn giữa cho cột UserID, Giờ vào, Giờ ra
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # Thiết lập hàng 5 làm header cố định
            ws.freeze_panes = ws['A6']  # Cố định từ hàng 5 trở xuống

            # Định dạng header (hàng 5)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Màu nền header
            header_font = Font(size=12, bold=True, color="FFFFFF")  # Định dạng font đậm và màu trắng

            for col in range(1, 8):  # Các cột A đến G
                cell = ws.cell(row=5, column=col)
                cell.fill = header_fill  # Áp dụng màu nền cho header
                cell.font = header_font  # Áp dụng font đậm và màu trắng
                cell.alignment = Alignment(horizontal='center', vertical='center')  # Căn giữa cho header

            # Điều chỉnh độ rộng cột
            column_widths = [15, 30, 30, 20, 15, 15, 30]  # Độ rộng cho từng cột
            for i, width in enumerate(column_widths):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width

            # Bật bộ lọc cho các cột từ hàng thứ 5
            ws.auto_filter.ref = f"A5:G{len(records) + 5}"  # Bật bộ lọc cho các cột từ A5 đến G

            # Lưu workbook
            wb.save(file_path)
            messagebox.showinfo("Thành công", "Đã xuất dữ liệu ra file Excel thành công.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xuất dữ liệu: {str(e)}")
        # Tải workbook hiện tại
        wb = load_workbook(file_path)

        # Cập nhật thông tin tác giả
        wb.properties.creator = "Face_Recognition v1.0.1"

        # Lưu lại file với thông tin tác giả mới
        wb.save(file_path)


# Khung quản lý dữ liệu điểm danh
attendance_frame = tk.Frame(root, width=1000, height=850)

# Nút chức năng quản lý dữ liệu điểm danh
attendance_function_frame = tk.Frame(attendance_frame)
attendance_function_frame.pack(side="top", anchor="center", padx=10, pady=10)

current_month = arrow.now().month # Lấy tháng hiện tại

month_label = tk.Label(attendance_function_frame, text="Chọn Tháng:")
month_label.pack(side="left", padx=5)

month_combo = ttk.Combobox(attendance_function_frame, values=list(range(1, 13)), width=5)
month_combo.set(current_month)  # Đặt giá trị mặc định là tháng hiện tại
month_combo.pack(side="left", padx=5)

filter_button = tk.Button(attendance_function_frame, text="Tải dữ liệu",image=load_btn, compound='left',padx=15,command=load_attention_data)
filter_button.pack(side="left", padx=25)

filter_button = tk.Button(attendance_function_frame, text="Lọc",image=filter_btn, compound='left',padx=15,command=open_filter_dialog)
filter_button.pack(side="left", padx=25)

export_button = tk.Button(attendance_function_frame, text="Xuất Excel",image=export_btn, compound='left',padx=15,command=export_to_excel)
export_button.pack(side="left", padx=25)

# Cột dữ liệu
attendance_columns = ('UserID', 'Tên','Vị trí', 'Ngày', 'Giờ vào', 'Giờ ra', 'Trạng thái')
attendance_treeview = ttk.Treeview(attendance_frame, columns=attendance_columns, show='headings')

for col in attendance_columns:
    attendance_treeview.heading(col, text=col)
# attendance_treeview.column('ID', width=30,anchor='center')# Độ rộng cho cột ID
attendance_treeview.column('UserID', width=40,anchor='center')
attendance_treeview.column('Tên', width=100)  # Độ rộng cho cột Tên# Độ rộng cho cột Vai trò
attendance_treeview.column('Vị trí', width=100)
attendance_treeview.column('Ngày', width=100,anchor='center')# Độ rộng cho cột Vai trò
attendance_treeview.column('Giờ vào', width=100,anchor='center')
attendance_treeview.column('Giờ ra', width=100,anchor='center')
attendance_treeview.column('Trạng thái', width=80)# Độ rộng cho cột Ghi chú
attendance_treeview.pack(fill='both', expand=True)



#DANH MỤC
def load_job_position():
    # Kết nối đến cơ sở dữ liệu
    server_name, database_name = load_db_config()
    if server_name and database_name:
        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )

            cursor = connection.cursor()
            cursor.execute("SELECT IdJob,job_position,timestart,timeend,note FROM job_position")  # Lấy 6 cột
            rows = cursor.fetchall()

            # Xóa dữ liệu cũ trong Treeview
            for item in catalog_treeview.get_children():
                catalog_treeview.delete(item)

            # Thêm dữ liệu mới vào Treeview
            for row in rows:
                start_time_str = row[2].split('.')[0] if row[2] else None  # Lấy phần trước dấu chấm
                end_time_str = row[3].split('.')[0] if row[3] else None  # Lấy phần trước dấu chấm
                
                start_time = (datetime.strptime(start_time_str, "%H:%M:%S").strftime("%H:%M") if start_time_str else None)
                end_time = (datetime.strptime(end_time_str, "%H:%M:%S").strftime("%H:%M") if end_time_str else None)
                
                
                catalog_treeview.insert('', 'end', values=(row[0], row[1], start_time,end_time, row[4]))
            # user_treeview.tag_configure('centered', anchor='center'),tags=('centered',)
            cursor.close()
            connection.close()
        except pyodbc.Error as ex:
            messagebox.showerror("Lỗi", "Lỗi tải dữ liệu người dùng!")
            # print("Error loading user data:", ex)
    else:
        messagebox.showerror("Lỗi", "Kiểm tra lại servername hoặc database!")
        # print("Server name or database name is missing.")


def add_job_position():
    # Kết nối tới cơ sở dữ liệu để lấy ID mới
    server_name, database_name = load_db_config()
    if not server_name or not database_name:
        messagebox.showerror("Lỗi", "Không thể tải cấu hình cơ sở dữ liệu.")
        return

    try:
        connection = pyodbc.connect(
            'DRIVER={SQL Server};' +
            f'Server={server_name};' +
            f'Database={database_name};' +
            'Trusted_Connection=yes;'
        )
        cursor = connection.cursor()

        # Tìm idJob cao nhất trong bảng
        cursor.execute("SELECT idJob FROM job_position ORDER BY idJob ASC")
        all_ids = [row[0] for row in cursor.fetchall()]

        # Tìm idJob nhỏ nhất còn thiếu hoặc tạo idJob mới nếu không thiếu
        idJob = 1
        for i in range(1, len(all_ids) + 1):
            if i != all_ids[i - 1]:
                idJob = i
                break
        else:
            # Nếu không có lỗ hổng trong dãy, đặt idJob mới là max_id + 1
            idJob = all_ids[-1] + 1 if all_ids else 1

        cursor.close()
        connection.close()
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể kết nối tới cơ sở dữ liệu: {e}")
        return

    def save_job_position():
        job_position = job_position_entry.get()
        start_hour = start_hour_entry.get()
        start_minute = start_minute_entry.get()
        end_hour = end_hour_entry.get()
        end_minute = end_minute_entry.get()
        note = note_entry.get()

        try:
            # Kiểm tra số ký tự tối đa
            if len(start_hour) > 2 or len(start_minute) > 2 or len(end_hour) > 2 or len(end_minute) > 2:
                messagebox.showerror("Lỗi", "Giờ và phút chỉ được nhập tối đa 2 ký tự.")
                return

            # Kiểm tra định dạng giờ và phút
            if not (start_hour.isdigit() and start_minute.isdigit() and 
                    end_hour.isdigit() and end_minute.isdigit()):
                messagebox.showerror("Lỗi", "Giờ và phút phải là số.")
                return
            
            # Kiểm tra giới hạn giá trị giờ và phút
            if not (0 <= int(start_hour) < 24 and 0 <= int(start_minute) < 60 and
                    0 <= int(end_hour) < 24 and 0 <= int(end_minute) < 60):
                messagebox.showerror("Lỗi", "Giờ phải từ 0 đến 23 và phút phải từ 0 đến 59.")
                return

            # Chuyển đổi giờ và phút thành kiểu time
            start_time = f"{int(start_hour):02}:{int(start_minute):02}"  # Chuyển thành định dạng "HH:MM"
            end_time = f"{int(end_hour):02}:{int(end_minute):02}"

            # Kiểm tra endtime lớn hơn starttime
            if end_time <= start_time:
                messagebox.showerror("Lỗi", "Thời gian kết thúc phải lớn hơn thời gian bắt đầu.")
                return

            # Kết nối tới cơ sở dữ liệu để thêm bản ghi
            server_name, database_name = load_db_config()
            if server_name and database_name:
                try:
                    connection = pyodbc.connect(
                        'DRIVER={SQL Server};' +
                        f'Server={server_name};' +
                        f'Database={database_name};' +
                        'Trusted_Connection=yes;'
                    )
                    cursor = connection.cursor()

                    # Thực hiện câu lệnh SQL để thêm bản ghi
                    sql_insert = """
                    INSERT INTO job_position (idJob, job_position, timestart, timeend, note)
                    VALUES (?, ?, ?, ?, ?)
                    """
                    cursor.execute(sql_insert, (idJob, job_position, start_time, end_time, note))
                    connection.commit()
                    cursor.close()
                    connection.close()
                    load_job_position()

                    messagebox.showinfo("Thành công", "Thêm chức vụ thành công!")
                    
                    add_job_position_dialog.destroy()

                except Exception as e:
                    messagebox.showerror("Lỗi", f"Đã xảy ra lỗi khi thêm chức vụ: {e}")
                    print(e)  # Hiển thị lỗi chi tiết trên console
            else:
                messagebox.showerror("Lỗi", "Không thể tải cấu hình cơ sở dữ liệu.")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Đã xảy ra lỗi: {e}")
            print(e)  # Hiển thị lỗi chi tiết trên console

    # Tạo dialog thêm chức vụ
    add_job_position_dialog = tk.Toplevel()
    add_job_position_dialog.title("Thêm Chức Vụ")
    add_job_position_dialog.geometry("400x300")
    center_dialog(add_job_position_dialog)
    add_job_position_dialog.grab_set()

    label_font = ('Arial', 12)
    entry_font = ('Arial', 12)

    # Hiển thị ID mới
    tk.Label(add_job_position_dialog, text=f"ID mới: {idJob}", font=label_font).place(relx=0.1, rely=0.05)

    tk.Label(add_job_position_dialog, text="Chức vụ:", font=label_font).place(relx=0.1, rely=0.2)
    job_position_entry = tk.Entry(add_job_position_dialog, font=entry_font)
    job_position_entry.place(relx=0.3, rely=0.2, width=230, height=30)

    tk.Label(add_job_position_dialog, text="Thời gian bắt đầu:", font=label_font).place(relx=0.1, rely=0.35)
    start_hour_entry = tk.Entry(add_job_position_dialog, font=entry_font, width=5)
    start_hour_entry.place(relx=0.45, rely=0.35, height=30)
    tk.Label(add_job_position_dialog, text=":", font=("Arial", 12, "bold")).place(relx=0.6, rely=0.35)
    start_minute_entry = tk.Entry(add_job_position_dialog, font=entry_font, width=5)
    start_minute_entry.place(relx=0.65, rely=0.35, height=30)

    tk.Label(add_job_position_dialog, text="Thời gian kết thúc:", font=label_font).place(relx=0.1, rely=0.5)
    end_hour_entry = tk.Entry(add_job_position_dialog, font=entry_font, width=5)
    end_hour_entry.place(relx=0.45, rely=0.5, height=30)
    tk.Label(add_job_position_dialog, text=":", font=("Arial", 12, "bold")).place(relx=0.6, rely=0.5)
    end_minute_entry = tk.Entry(add_job_position_dialog, font=entry_font, width=5)
    end_minute_entry.place(relx=0.65, rely=0.5, height=30)

    tk.Label(add_job_position_dialog, text="Ghi chú:", font=label_font).place(relx=0.1, rely=0.65)
    note_entry = tk.Entry(add_job_position_dialog,font=entry_font)
    note_entry.place(relx=0.3, rely=0.65, width=230, height=30)

    add_button = tk.Button(add_job_position_dialog, text="Thêm Mới", command=save_job_position, font=("Arial", 12), bg="green")
    add_button.place(relx=0.4, rely=0.8, width=100, height=40)




def edit_job_position():
    # Kiểm tra xem có chọn một bản ghi nào chưa
    selected_item = catalog_treeview.selection()
    if not selected_item:
        messagebox.showerror("Lỗi", "Vui lòng chọn một bản ghi để sửa!")
        # print("Vui lòng chọn một bản ghi để sửa.")
        return

    # Lấy ID của bản ghi đã chọn
    selected_id = catalog_treeview.item(selected_item, 'values')[0]

    # Kết nối đến cơ sở dữ liệu và lấy dữ liệu của bản ghi đã chọn
    server_name, database_name = load_db_config()
    if server_name and database_name:
        try:
            connection = pyodbc.connect(
                'DRIVER={SQL Server};' +
                f'Server={server_name};' +
                f'Database={database_name};' +
                'Trusted_Connection=True'
            )
            cursor = connection.cursor()
            cursor.execute("SELECT idjob, job_position, timestart, timeend, note FROM job_position WHERE idjob = ?", (selected_id,))
            row = cursor.fetchone()
            if row:
                # In ra giá trị để kiểm tra định dạng
                print(f"Start time: {row[2]}, End time: {row[3]}")

                # Chuyển đổi thời gian bắt đầu và kết thúc
                start_time = row[2] or None
                end_time = row[3] or None
                
                # Xử lý thời gian
                if start_time:
                    start_time = start_time.split('.')[0]  # Cắt phần sau dấu chấm
                    start_time = start_time.split(':')[:2]  # Chỉ lấy giờ và phút
                    start_hours, start_minutes = start_time[0], start_time[1]
                else:
                    start_hours, start_minutes = '00', '00'  # Mặc định nếu không có

                if end_time:
                    end_time = end_time.split('.')[0]  # Cắt phần sau dấu chấm
                    end_time = end_time.split(':')[:2]  # Chỉ lấy giờ và phút
                    end_hours, end_minutes = end_time[0], end_time[1]
                else:
                    end_hours, end_minutes = '00', '00'  # Mặc định nếu không có

                # Mở dialog với thông tin đã lấy
                dialog = customtkinter.CTkToplevel(root)
                dialog.title("Sửa Vị Trí Công Việc")
                center_dialog(dialog)
                dialog.geometry("400x450")

                dialog.grab_set()  # Khóa dialog lại

                # Hiển thị thông tin và các trường chỉnh sửa
                customtkinter.CTkLabel(dialog, text=f"ID: {row[0]}").place(relx=0.1, rely=0.05)

                customtkinter.CTkLabel(dialog, text="Tên Vị Trí:").place(relx=0.1, rely=0.12)
                position_entry = customtkinter.CTkEntry(dialog, width=310)
                position_entry.place(relx=0.1, rely=0.17)
                position_entry.insert(0, row[1] or "")

                customtkinter.CTkLabel(dialog, text="Thời Gian Bắt Đầu:").place(relx=0.1, rely=0.27)
                starttime_hour_entry = customtkinter.CTkEntry(dialog, width=130)
                starttime_hour_entry.place(relx=0.1, rely=0.32)
                starttime_hour_entry.insert(0, start_hours)

                starttime_minute_entry = customtkinter.CTkEntry(dialog, width=130)
                starttime_minute_entry.place(relx=0.55, rely=0.32)
                starttime_minute_entry.insert(0, start_minutes)

                customtkinter.CTkLabel(dialog, text="Thời Gian Kết Thúc:").place(relx=0.1, rely=0.42)
                endtime_hour_entry = customtkinter.CTkEntry(dialog, width=130)
                endtime_hour_entry.place(relx=0.1, rely=0.47)
                endtime_hour_entry.insert(0, end_hours)

                endtime_minute_entry = customtkinter.CTkEntry(dialog, width=130)
                endtime_minute_entry.place(relx=0.55, rely=0.47)
                endtime_minute_entry.insert(0, end_minutes)

                # Hiển thị ghi chú
                customtkinter.CTkLabel(dialog, text="Ghi Chú:").place(relx=0.1, rely=0.57)
                note_entry = customtkinter.CTkEntry(dialog, width=310)
                note_entry.place(relx=0.1, rely=0.62)
                note_entry.insert(0, row[4] or "")

                # Label thông báo kết quả
                result_label = customtkinter.CTkLabel(dialog, text="")
                result_label.place(relx=0.1, rely=0.85)

                # Hàm lưu thông tin đã chỉnh sửa
                def save_changes():
                    new_position = position_entry.get()
                    new_start_hours = starttime_hour_entry.get()
                    new_start_minutes = starttime_minute_entry.get()
                    new_end_hours = endtime_hour_entry.get()
                    new_end_minutes = endtime_minute_entry.get()
                    new_note = note_entry.get()

                    # Kiểm tra tính hợp lệ của thời gian
                    try:
                        start_hours = int(new_start_hours)
                        start_minutes = int(new_start_minutes)
                        end_hours = int(new_end_hours)
                        end_minutes = int(new_end_minutes)

                        # Kiểm tra xem các giá trị thời gian có hợp lệ không
                        if not (0 <= start_hours < 24 and 0 <= start_minutes < 60 and 0 <= end_hours < 24 and 0 <= end_minutes < 60):
                            result_label.configure(text="Thời gian không hợp lệ. Vui lòng kiểm tra lại.")
                            return
                        
                        # Kiểm tra xem endtime có lớn hơn starttime không
                        if (end_hours < start_hours) or (end_hours == start_hours and end_minutes <= start_minutes):
                            result_label.configure(text="Thời gian kết thúc phải lớn hơn thời gian bắt đầu.")
                            return

                        # Chuyển đổi thời gian sang định dạng phù hợp
                        new_starttime = f"{start_hours:02}:{start_minutes:02}:00"  # Thêm giây
                        new_endtime = f"{end_hours:02}:{end_minutes:02}:00"  # Thêm giây

                    except ValueError:
                        result_label.configure(text="Định dạng thời gian không hợp lệ. Vui lòng nhập lại.")
                        return  # Ngưng thực hiện nếu không thể chuyển đổi thành số

                    # Nếu thời gian hợp lệ, thực hiện lưu thay đổi
                    try:
                        cursor.execute(
                            "UPDATE job_position SET job_position = ?, timestart = ?, timeend = ?, note = ? WHERE idjob = ?",
                            (new_position, new_starttime, new_endtime, new_note, row[0])
                        )
                        connection.commit()
                        cursor.close()
                        connection.close()
                        result_label.configure(text="Thay đổi thông tin thành công!")
                        load_job_position()  # Tải lại dữ liệu vị trí công việc
                    except pyodbc.Error as ex:
                        # print("Error saving changes:", ex)
                        result_label.configure(text="Lỗi khi lưu thay đổi.")

                # Nút Lưu để cập nhật thông tin
                save_button = customtkinter.CTkButton(dialog, text="Lưu", command=save_changes, fg_color="green")
                save_button.place(relx=0.35, rely=0.9)

                
        except pyodbc.Error as ex:
            messagebox.showerror("Lỗi", "Lỗi tải vị trí công việc để sửa!"+ex)
            # print("Error loading job position data for edit:", ex)
    else:
        messagebox.showerror("Lỗi", "Kiểm tra lại servername hoặc database!")
        # print("Server name or database name is missing.")

def delete_job_position():
    # Kiểm tra xem có chọn một bản ghi nào chưa
    selected_item = catalog_treeview.selection()
    if not selected_item:
        messagebox.showerror("Lỗi", "Vui lòng chọn một bản ghi để xoá!")
        # print("Vui lòng chọn một bản ghi để xóa.")
        return

    # Lấy ID của bản ghi đã chọn
    selected_id = catalog_treeview.item(selected_item, 'values')[0]

    # Tạo dialog xác nhận
    confirm_dialog = customtkinter.CTkToplevel(root)
    confirm_dialog.title("Xác Nhận Xóa")
    confirm_dialog.geometry("300x150")
    center_dialog(confirm_dialog)
    confirm_dialog.grab_set()  # Khóa dialog lại

    # Thông báo
    customtkinter.CTkLabel(confirm_dialog, text=f"Bạn có muốn xóa bản ghi này không?").pack(pady=20)

    def confirm_delete():
        server_name, database_name = load_db_config()
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )
                cursor = connection.cursor()

                # Thực hiện xóa bản ghi
                cursor.execute("DELETE FROM job_position WHERE idjob = ?", (selected_id,))
                connection.commit()
                
                cursor.close()
                connection.close()

                # Thông báo xóa thành công và cập nhật Treeview
                print(f"Bản ghi ID: {selected_id} đã được xóa thành công.")
                load_job_position()  # Tải lại dữ liệu vị trí công việc
                confirm_dialog.destroy()  # Đóng dialog xác nhận
            except pyodbc.Error as ex:
                messagebox.showerror("Lỗi", "Lỗi khi xoá bản ghi!"+ex)
                # print("Lỗi khi xóa bản ghi:", ex)
        else:
            messagebox.showerror("Lỗi", "Kiểm tra lại servername hoặc database!")
            # print("Server name or database name is missing.")

    def cancel_delete():
        confirm_dialog.destroy()  # Đóng dialog nếu người dùng hủy

    # Nút xác nhận
    confirm_button = customtkinter.CTkButton(confirm_dialog, text="Xác Nhận", command=confirm_delete, fg_color="green",width=100)
    confirm_button.pack(side=tk.LEFT, padx=(40,5), pady=10)

    # Nút hủy
    cancel_button = customtkinter.CTkButton(confirm_dialog, text="Hủy", command=cancel_delete, fg_color="red",width=100)
    cancel_button.pack(side=tk.RIGHT, padx=(5,40), pady=10)




def search_job_position():
    # Tạo dialog tìm kiếm
    search_dialog = customtkinter.CTkToplevel(root)
    search_dialog.title("Tìm Kiếm Vị Trí Công Việc")
    search_dialog.geometry("400x300")
    search_dialog.grab_set()  # Khóa dialog lại
    center_dialog(search_dialog)

    # Nhập tên vị trí
    customtkinter.CTkLabel(search_dialog, text="Tên Vị Trí:").place(relx=0.1, rely=0.1)
    job_position_entry = customtkinter.CTkEntry(search_dialog, width=230)
    job_position_entry.place(relx=0.3, rely=0.1)

    # Nhập thời gian bắt đầu
    customtkinter.CTkLabel(search_dialog, text="Thời Gian Bắt Đầu:").place(relx=0.1, rely=0.32)
    start_hour_entry = customtkinter.CTkEntry(search_dialog, width=40)
    start_hour_entry.place(relx=0.5, rely=0.32)
    customtkinter.CTkLabel(search_dialog, text=":").place(relx=0.62, rely=0.32)
    start_minute_entry = customtkinter.CTkEntry(search_dialog, width=40)
    start_minute_entry.insert(0, "00")  # Đặt placeholder là "00"
    start_minute_entry.place(relx=0.65, rely=0.32)

    # Nhập thời gian kết thúc
    customtkinter.CTkLabel(search_dialog, text="Thời Gian Kết Thúc:").place(relx=0.1, rely=0.54)
    end_hour_entry = customtkinter.CTkEntry(search_dialog, width=40)
    end_hour_entry.place(relx=0.5, rely=0.54)
    customtkinter.CTkLabel(search_dialog, text=":").place(relx=0.62, rely=0.54)
    end_minute_entry = customtkinter.CTkEntry(search_dialog, width=40)
    end_minute_entry.insert(0, "00")  # Đặt placeholder là "00"
    end_minute_entry.place(relx=0.65, rely=0.54)

    def execute_search():
        # Lấy giá trị từ các ô nhập liệu
        position_name = job_position_entry.get().strip()
        start_hour = start_hour_entry.get().strip()
        start_minute = start_minute_entry.get().strip()   # Mặc định là "00" nếu không nhập phút
        end_hour = end_hour_entry.get().strip()
        end_minute = end_minute_entry.get().strip()
        # or "00"  # Mặc định là "00" nếu không nhập phút

        # Kiểm tra nếu tất cả các trường đều trống
        if not position_name and not start_hour and not end_hour:
            messagebox.showerror("Lỗi", "Vui lòng nhập ít nhất một trường tìm kiếm!")
            # print("Vui lòng nhập ít nhất một trường tìm kiếm.")
            return

        # Kiểm tra định dạng giờ và phút
        if (start_hour and not start_hour.isdigit()) or (start_minute and not start_minute.isdigit()):
            messagebox.showerror("Lỗi", "Giờ và phút bắt đầu phải là số!")
            # print("Giờ và phút bắt đầu phải là số.")
            return
        if (end_hour and not end_hour.isdigit()) or (end_minute and not end_minute.isdigit()):
            messagebox.showerror("Lỗi", "Giờ và phút kết thúc phải là số!")
            # print("Giờ và phút kết thúc phải là số.")
            return

        # Chuyển đổi giờ và phút thành định dạng HH:MM
        start_time = f"{start_hour.zfill(2)}:{start_minute.zfill(2)}" if start_hour else ""
        end_time = f"{end_hour.zfill(2)}:{end_minute.zfill(2)}" if end_hour else ""

        # Chuẩn bị câu truy vấn SQL
        query = "SELECT * FROM job_position WHERE 1=1"
        parameters = []

        if position_name:
            query += " AND job_position LIKE ?"
            parameters.append(f"%{position_name}%")

        if start_time:
            query += " AND timestart >= ?"
            parameters.append(start_time)

        if end_time:
            query += " AND timeend >= ?"
            parameters.append(end_time)

        server_name, database_name = load_db_config()
        if server_name and database_name:
            try:
                connection = pyodbc.connect(
                    'DRIVER={SQL Server};' +
                    f'Server={server_name};' +
                    f'Database={database_name};' +
                    'Trusted_Connection=True'
                )
                cursor = connection.cursor()

                # Thực hiện truy vấn
                cursor.execute(query, parameters)
                results = cursor.fetchall()

                # Xóa dữ liệu cũ trong Treeview
                catalog_treeview.delete(*catalog_treeview.get_children())

                # Hiển thị kết quả trên Treeview
                for row in results:
                    start_time_str = row[2]
                    end_time_str = row[3]

                    # Chuyển định dạng HH:MM:SS thành HH:MM
                    start_time_formatted = start_time_str.split('.')[0][:-3] if start_time_str else None
                    end_time_formatted = end_time_str.split('.')[0][:-3] if end_time_str else None

                    catalog_treeview.insert('', 'end', values=(row[0], row[1], start_time_formatted, end_time_formatted, row[4]))

                cursor.close()
                connection.close()
                
                            # Hiển thị thông báo số bản ghi tìm thấy
                num_records = len(results)
                messagebox.showinfo("Kết quả tìm kiếm", f"Tìm thấy {num_records} bản ghi!")
                
                # messagebox.showinfo("Thành công", "Tìm kiếm thành công!")
                # print("Tìm kiếm thành công.")
                search_dialog.destroy()  # Đóng dialog tìm kiếm
            except pyodbc.Error as ex:
                messagebox.showerror("Lỗi", "Lỗi khi tìm kiếm"+ex)
                # print("Lỗi khi tìm kiếm:", ex)
        else:
            messagebox.showerror("Lỗi", "Kiểm tra lại servername hoặc database!")
            # print("Server name or database name is missing.")

    # Nút tìm kiếm
    search_button = customtkinter.CTkButton(search_dialog, text="Tìm Kiếm", command=execute_search, fg_color="blue", width=150)
    search_button.place(relx=0.1, rely=0.8)

    # Nút hủy
    cancel_button = customtkinter.CTkButton(search_dialog, text="Hủy", command=search_dialog.destroy, fg_color="red", width=150)
    cancel_button.place(relx=0.5, rely=0.8)








# Khung danh muc
catalog_frame = tk.Frame(root, width=1000, height=850)

# Nút chức năng danh muc
catalog_function_frame = tk.Frame(catalog_frame)
catalog_function_frame.pack(side="top", anchor="center", padx=10, pady=10)

filter_button = tk.Button(catalog_function_frame, text="Tải dữ liệu",image=load_btn, compound='left',padx=12,command=load_job_position)
filter_button.pack(side="left", padx=15)

filter_button = tk.Button(catalog_function_frame, text="Thêm",image=add_btn, compound='left',padx=12,command=add_job_position)
filter_button.pack(side="left", padx=15)

export_button = tk.Button(catalog_function_frame, text="Sửa",image=edit_btn, compound='left',padx=12,command=edit_job_position)
export_button.pack(side="left", padx=15)

export_button = tk.Button(catalog_function_frame, text="Xoá",image=delete_btn, compound='left',padx=12,command=delete_job_position)
export_button.pack(side="left", padx=15)

export_button = tk.Button(catalog_function_frame, text="Tìm kiếm",image=search_btn, compound='left',padx=15,command=search_job_position)
export_button.pack(side="left", padx=15)

# Cột dữ liệu danh muc
catalog_columns = ('ID','Vị trí','Bắt đầu','Kết thúc','Ghi chú')
catalog_treeview = ttk.Treeview(catalog_frame, columns=catalog_columns, show='headings')
for col in catalog_columns:
    catalog_treeview.heading(col, text=col)
catalog_treeview.column('ID', width=10,anchor='center')  # Độ rộng cho cột Vai trò
catalog_treeview.column('Vị trí', width=80)  # Độ rộng cho cột Vai trò
catalog_treeview.column('Bắt đầu', width=50,anchor='center')  # Độ rộng cho cột Vai trò
catalog_treeview.column('Kết thúc', width=50,anchor='center')  # Độ rộng cho cột Vai trò
catalog_treeview.column('Ghi chú', width=200)  # Độ rộng cho cột Vai trò
catalog_treeview.pack(fill='both', expand=True)



# Cập nhật thời gian
update_time()



root.mainloop()
